import pandas as pd  # type: ignore
import plotly.express as px  # type: ignore
import plotly.graph_objects as go  # type: ignore
from datetime import datetime, date
from pathlib import Path
import shutil
import json
from typing import Optional
import re
import urllib.parse
import urllib.request
from io import BytesIO
from openpyxl import load_workbook  # type: ignore
import sqlite3
import logging
import streamlit as st  # type: ignore
import socket

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
REFERENCE_DATA_DIR = BASE_DIR / "reference_data"

TIKTOK_FILENAME = "TikTok 90 days history views - Sheet1.csv"
TIKTOK_DATA_PATH = DATA_DIR / TIKTOK_FILENAME
TIKTOK_FALLBACK_PATH = REFERENCE_DATA_DIR / TIKTOK_FILENAME

# Current mode file paths
CURRENT_VIEWS_FILENAME = "current_views.csv"
CURRENT_VIEWS_PATH = DATA_DIR / CURRENT_VIEWS_FILENAME
BSR_MANUAL_ENTRIES_PATH = DATA_DIR / "manual_bsr_entries.json"
MANUAL_TIKTOK_LINKS_PATH = DATA_DIR / "manual_tiktok_links.json"
DB_PATH = DATA_DIR / "app.db"

# Recent 7-day exports
RECENT_CORE_FILENAME = "tiktok_core_latest.csv"
RECENT_CORE_DATA_PATH = DATA_DIR / RECENT_CORE_FILENAME
VIDEO_DETAILS_FILENAME = "tiktok_video_details_latest.csv"
VIDEO_DETAILS_DATA_PATH = DATA_DIR / VIDEO_DETAILS_FILENAME

# Brand-specific data paths
def get_brand_core_path(brand: str) -> Path:
    """Get core data path for a specific brand."""
    brand_normalized = brand.lower().replace(" ", "_")
    return DATA_DIR / f"tiktok_core_{brand_normalized}.csv"

def get_brand_video_details_path(brand: str) -> Path:
    """Get video details path for a specific brand."""
    brand_normalized = brand.lower().replace(" ", "_")
    return DATA_DIR / f"tiktok_video_details_{brand_normalized}.csv"


def _ensure_dataset(csv_path: Path) -> Path:
    csv_path = Path(csv_path)
    if csv_path.exists():
        return csv_path

    if TIKTOK_FALLBACK_PATH.exists():
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(TIKTOK_FALLBACK_PATH, csv_path)
        return csv_path

    return csv_path


def _parse_tiktok_dates(raw_dates: pd.Series) -> pd.Series:
    raw = raw_dates.astype(str).str.strip()

    parsed = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")

    def parse_with_format(fmt: str):
        mask = parsed.isna()
        if not mask.any():
            return
        parsed.loc[mask] = pd.to_datetime(raw.loc[mask], format=fmt, errors="coerce")

    # Try known formats explicitly to avoid pandas inference warnings
    parse_with_format("%Y-%m-%d %H:%M:%S")
    parse_with_format("%Y-%m-%d")
    parse_with_format("%m/%d/%Y")

    # Handle month/day values by appending the current year
    mask = parsed.isna()
    if mask.any():
        year_appended = raw.loc[mask] + f"/{datetime.now().year}"
        parsed.loc[mask] = pd.to_datetime(year_appended, format="%m/%d/%Y", errors="coerce")

    # As a last resort, parse bare month/day and set the current year
    mask = parsed.isna()
    if mask.any():
        month_day = pd.to_datetime(raw.loc[mask], format="%m/%d", errors="coerce")
        parsed.loc[mask] = month_day.apply(
            lambda dt: dt.replace(year=datetime.now().year) if pd.notna(dt) else dt
        )

    return parsed


def load_views_bsr_data(csv_path: Path = TIKTOK_DATA_PATH) -> pd.DataFrame:
    csv_path = _ensure_dataset(csv_path)
    if not csv_path.exists():
        return pd.DataFrame()

    df_raw = pd.read_csv(csv_path, header=1)
    df_raw.columns = df_raw.columns.str.strip()

    sum_of_views_col = None
    for col in df_raw.columns:
        if col.strip().lower() == "sum of views":
            sum_of_views_col = col
            break

    if sum_of_views_col:
        total_views = pd.to_numeric(
            df_raw[sum_of_views_col].astype(str).str.replace(r"[^\d.-]", "", regex=True),
            errors="coerce",
        )
    else:
        view_cols = [col for col in df_raw.columns if "view" in col.lower()]
        numeric_cols = df_raw[view_cols].apply(
            lambda col: pd.to_numeric(
                col.astype(str).str.replace(r"[^\d.-]", "", regex=True), errors="coerce"
            )
        )
        total_views = numeric_cols.sum(axis=1, skipna=True)

    df = pd.DataFrame()
    df["date"] = _parse_tiktok_dates(df_raw["date"])
    df["total_views"] = total_views
    df["BSR Amazon"] = pd.to_numeric(
        df_raw.get("BSR Amazon", pd.Series(dtype=float)),
        errors="coerce"
    )

    exclude_days = {"10-03", "10-04", "10-05", "10-06"}
    if not df["date"].isna().all():
        df = df[~df["date"].dt.strftime("%m-%d").isin(exclude_days)]

    df = df.dropna(subset=["date", "BSR Amazon"])
    df = df[df["total_views"].notna()]
    # Exclude days with zero or negative views
    df = df[df["total_views"] > 0]
    df = df.sort_values("date")
    return df[["date", "total_views", "BSR Amazon"]]


@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_recent_core_data(csv_path: Path = RECENT_CORE_DATA_PATH) -> pd.DataFrame:
    """Load the 7-day core export (Original/Repost/Total metrics).
    
    Note: Cache key includes the full path string to ensure brand-specific data
    is cached separately.
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        return pd.DataFrame()

    try:
        # Use the full absolute path as part of cache key to ensure brand separation
        df = pd.read_csv(csv_path)
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    df.columns = df.columns.str.strip()
    date_col = "Date" if "Date" in df.columns else df.columns[0]
    df["Date"] = _parse_tiktok_dates(df[date_col])
    numeric_cols = [col for col in df.columns if col != "Date"]
    if numeric_cols:
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df.sort_values("Date")
    return df.reset_index(drop=True)


def _normalize_current_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["date", "total_views", "BSR Amazon"])
    normalized = pd.DataFrame()
    normalized["date"] = pd.to_datetime(df.get("date"), errors="coerce")
    normalized["total_views"] = pd.to_numeric(df.get("total_views"), errors="coerce")
    normalized["BSR Amazon"] = pd.to_numeric(df.get("BSR Amazon"), errors="coerce")
    normalized = normalized.dropna(subset=["date"])
    return normalized


def _core_to_current_frame(core_df: pd.DataFrame | None) -> pd.DataFrame:
    if core_df is None or core_df.empty:
        return pd.DataFrame(columns=["date", "total_views", "BSR Amazon"])

    df = core_df.copy()
    date_col = "Date" if "Date" in df.columns else df.columns[0]
    df[date_col] = _parse_tiktok_dates(df[date_col])
    df = df.sort_values(date_col).reset_index(drop=True)

    if "Total Views" in df.columns:
        # Total Views are cumulative - convert to daily differences
        # Use Original Views + Repost Views differences for more accurate daily calculation
        if "Original Views" in df.columns and "Repost Views" in df.columns:
            original_views = pd.to_numeric(df["Original Views"], errors="coerce")
            repost_views = pd.to_numeric(df["Repost Views"], errors="coerce")
            # Calculate daily differences for each
            original_daily = original_views.diff().fillna(0)
            repost_daily = repost_views.diff().fillna(0)
            # Sum to get total daily view change
            daily_views = original_daily + repost_daily
        else:
            # Fallback to Total Views difference
            cumulative_views = pd.to_numeric(df["Total Views"], errors="coerce")
            daily_views = cumulative_views.diff().fillna(0)
        views_series = daily_views.fillna(0)
    else:
        view_cols = [
            col for col in df.columns if isinstance(col, str) and "view" in col.lower()
        ]
        if view_cols:
            cumulative_views = (
                df[view_cols]
                .apply(pd.to_numeric, errors="coerce")
                .sum(axis=1, skipna=True)
            )
            # Convert cumulative to daily differences
            daily_views = cumulative_views.diff().fillna(cumulative_views.iloc[0] if len(cumulative_views) > 0 and pd.notna(cumulative_views.iloc[0]) else 0)
            views_series = daily_views
        else:
            views_series = pd.Series(
                [float("nan")] * len(df), index=df.index, dtype="float64"
            )

    bsr_col = next(
        (col for col in df.columns if isinstance(col, str) and "bsr" in col.lower()),
        None,
    )
    if bsr_col:
        bsr_series = pd.to_numeric(df[bsr_col], errors="coerce")
    else:
        bsr_series = pd.Series(
            [float("nan")] * len(df), index=df.index, dtype="float64"
        )

    current = pd.DataFrame(
        {
            "date": pd.to_datetime(df[date_col], errors="coerce"),
            "total_views": views_series,
            "BSR Amazon": bsr_series,
        }
    )
    current = current.dropna(subset=["date"])
    # Filter out days with zero views (e.g., first day when calculating differences)
    current = current[current["total_views"] > 0]
    return current


def _apply_manual_bsr(
    df: pd.DataFrame, manual_entries: list[dict] | None
) -> pd.DataFrame:
    manual_entries = manual_entries or []
    result = df.copy()
    if "BSR Amazon" not in result.columns:
        result["BSR Amazon"] = pd.NA
    if result.empty and not manual_entries:
        return result

    for entry in manual_entries:
        entry_date = pd.to_datetime(entry.get("date"), errors="coerce")
        if pd.isna(entry_date):
            continue
        entry_value = entry.get("bsr")
        entry_value = float(entry_value) if entry_value is not None else None
        mask = result["date"].dt.normalize() == entry_date.normalize()
        if mask.any():
            result.loc[mask, "BSR Amazon"] = entry_value
        else:
            # Add new row for BSR-only entry (no view data for this date)
            new_row = pd.DataFrame(
                {
                    "date": [entry_date],
                    "total_views": [0.0],  # Use 0.0 instead of 0 to ensure it's numeric
                    "BSR Amazon": [entry_value],
                }
            )
            result = pd.concat([result, new_row], ignore_index=True)
    return result


@st.cache_data(ttl=10)  # Cache for 10 seconds (shorter to pick up new BSR entries faster)
def create_current_dataset(
    primary_df: pd.DataFrame | None,
    core_df: pd.DataFrame | None,
    manual_entries: list[dict] | None,
) -> pd.DataFrame:
    frames = []
    normalized_primary = _normalize_current_df(primary_df)
    if not normalized_primary.empty:
        frames.append(normalized_primary)

    core_frame = _core_to_current_frame(core_df)
    if not core_frame.empty:
        frames.append(core_frame)

    if frames:
        combined = pd.concat(frames, ignore_index=True)
        combined = combined.sort_values("date")
        # When dropping duplicates, prefer rows with BSR data
        # Sort by BSR (NaN last) then keep last to prioritize BSR entries
        combined = combined.sort_values("BSR Amazon", na_position='last')
        combined = combined.drop_duplicates(subset="date", keep="last")
        # Re-sort by date
        combined = combined.sort_values("date")
    else:
        combined = pd.DataFrame(columns=["date", "total_views", "BSR Amazon"])

    combined = _apply_manual_bsr(combined, manual_entries)

    # Don't filter by core date range if we have manual BSR entries outside that range
    # Manual BSR entries should always be included
    if core_df is not None and not core_df.empty and "Date" in core_df.columns:
        core_dates = pd.to_datetime(core_df["Date"], errors="coerce").dropna()
        if not core_dates.empty:
            min_core = core_dates.min()
            max_core = core_dates.max()
            # Get dates from manual entries to ensure they're not filtered out
            manual_entry_dates = set()
            if manual_entries:
                for entry in manual_entries:
                    entry_date = pd.to_datetime(entry.get("date"), errors="coerce")
                    if pd.notna(entry_date):
                        manual_entry_dates.add(entry_date.normalize())
            
            # Filter, but keep manual BSR entry dates (even if they have 0 views)
            date_mask = (combined["date"] >= min_core) & (combined["date"] <= max_core)
            manual_mask = combined["date"].dt.normalize().isin(manual_entry_dates)
            combined = combined[date_mask | manual_mask]
    
    # Don't filter out rows with BSR data even if views are 0
    # Only filter out rows with no views AND no BSR
    combined = combined[
        (combined["total_views"] > 0) | (combined["BSR Amazon"].notna())
    ]

    combined = combined.sort_values("date")
    return combined.reset_index(drop=True)


def create_views_vs_bsr_chart(df: pd.DataFrame) -> go.Figure | None:
    if df.empty:
        return None

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=df["date"],
            y=df["total_views"],
            name="TikTok Views",
            mode="lines+markers",
            line=dict(color="#1f77b4"),
            hovertemplate="Date: %{x|%b %d, %Y}<br>Views: %{y:,.0f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=df["date"],
            y=df["BSR Amazon"],
            name="Amazon BSR",
            mode="lines+markers",
            line=dict(color="#ff7f0e"),
            yaxis="y2",
            hovertemplate="Date: %{x|%b %d, %Y}<br>BSR: %{y:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        title="TikTok Views vs. Amazon BSR",
        xaxis=dict(title="Date"),
        yaxis=dict(title="TikTok Views"),
        yaxis2=dict(
            title="Amazon BSR",
            overlaying="y",
            side="right",
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
        height=450,
        template="plotly_white",
    )
    return fig


def create_views_change_vs_bsr_chart(daily_df: pd.DataFrame) -> go.Figure | None:
    """Plot day-over-day view changes against BSR."""
    if daily_df.empty or "views_change" not in daily_df.columns:
        return None

    df = daily_df.copy()
    if "date" not in df.columns:
        return None

    df["views_change"] = pd.to_numeric(df["views_change"], errors="coerce")
    df["BSR Amazon"] = pd.to_numeric(df["BSR Amazon"], errors="coerce")
    df = df.dropna(subset=["date", "views_change", "BSR Amazon"])
    if df.empty:
        return None

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=df["date"],
            y=df["views_change"],
            name="Δ TikTok Views",
            mode="lines+markers",
            line=dict(color="#1f77b4"),
            hovertemplate="Date: %{x|%b %d, %Y}<br>Δ Views: %{y:,.0f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=df["date"],
            y=df["BSR Amazon"],
            name="Amazon BSR",
            mode="lines+markers",
            line=dict(color="#ff7f0e"),
            yaxis="y2",
            hovertemplate="Date: %{x|%b %d, %Y}<br>BSR: %{y:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        title="Day-over-day TikTok Views vs. Amazon BSR",
        xaxis=dict(title="Date"),
        yaxis=dict(title="Δ TikTok Views"),
        yaxis2=dict(
            title="Amazon BSR",
            overlaying="y",
            side="right",
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
        height=450,
        template="plotly_white",
    )
    return fig


def create_views_line_chart(df: pd.DataFrame) -> go.Figure | None:
    if df.empty:
        return None
    fig = px.line(
        df,
        x="date",
        y="total_views",
        title="Daily View Changes",
        markers=True,
    )
    fig.update_layout(xaxis_title="Date", yaxis_title="Daily View Changes", template="plotly_white")
    return fig


def create_bsr_line_chart(df: pd.DataFrame) -> go.Figure | None:
    if df.empty:
        return None
    fig = px.line(
        df,
        x="date",
        y="BSR Amazon",
        title="Amazon BSR Over Time",
        markers=True,
    )
    fig.update_layout(
        xaxis_title="Date",
        yaxis_title="Amazon BSR",
        template="plotly_white",
    )
    return fig


def create_core_views_chart(df: pd.DataFrame) -> go.Figure | None:
    """Compare Original/Repost/Total views for recent core export."""
    if df.empty or "Date" not in df.columns:
        return None

    value_cols = [col for col in ["Original Views", "Repost Views", "Total Views"] if col in df.columns]
    if not value_cols:
        return None

    chart_df = df.melt(
        id_vars="Date",
        value_vars=value_cols,
        var_name="Series",
        value_name="Views",
    )
    fig = px.line(
        chart_df,
        x="Date",
        y="Views",
        color="Series",
        markers=True,
        title="Original vs Repost vs Total Views",
    )
    fig.update_layout(template="plotly_white", xaxis_title="Date", yaxis_title="Views")
    return fig


def create_core_engagement_chart(df: pd.DataFrame) -> go.Figure | None:
    """Stacked area chart for likes/comments/shares."""
    if df.empty or "Date" not in df.columns:
        return None

    value_cols = [col for col in ["Total Likes", "Total Comments", "Total Shares"] if col in df.columns]
    if not value_cols:
        return None

    chart_df = df.melt(
        id_vars="Date",
        value_vars=value_cols,
        var_name="Engagement",
        value_name="Count",
    )
    fig = px.area(
        chart_df,
        x="Date",
        y="Count",
        color="Engagement",
        title="Engagement Trends",
    )
    fig.update_layout(template="plotly_white", xaxis_title="Date", yaxis_title="Count")
    return fig


def create_repost_views_chart(df: pd.DataFrame) -> go.Figure | None:
    """Single-series line chart for daily change in repost views."""
    if df.empty or "Date" not in df.columns or "Repost Views" not in df.columns:
        return None

    chart_df = df[["Date", "Repost Views"]].copy()
    chart_df["Date"] = pd.to_datetime(chart_df["Date"], errors="coerce")
    chart_df = chart_df.dropna(subset=["Date"])
    chart_df = chart_df.sort_values("Date")
    chart_df["Repost Views"] = chart_df["Repost Views"].diff().fillna(0)
    if chart_df.empty:
        return None

    fig = px.line(
        chart_df,
        x="Date",
        y="Repost Views",
        title="Daily Reposted Views",
        markers=True,
    )
    fig.update_layout(template="plotly_white", xaxis_title="Date", yaxis_title="Daily Repost Views")
    return fig


def create_video_growth_scatter(summary_df: pd.DataFrame) -> go.Figure | None:
    """Scatter plot showing views delta vs. latest views for standout videos."""
    if summary_df.empty:
        return None

    fig = px.scatter(
        summary_df,
        x="views_delta",
        y="latest_views",
        color="video_type",
        size="latest_likes",
        hover_data={
            "video_id": True,
            "views_delta": ":,",
            "latest_views": ":,",
            "latest_likes": ":,",
            "latest_comments": ":,",
            "latest_shares": ":,",
            "video_url": True,
        },
        title="Video Growth vs Latest Views",
    )
    fig.update_layout(
        template="plotly_white",
        xaxis_title="7-Day Views Change",
        yaxis_title="Latest Views",
    )
    return fig


@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_current_views_data(csv_path: Path = CURRENT_VIEWS_PATH) -> pd.DataFrame:
    """Load current mode views data from daily updated file.
    
    Supports multiple CSV formats:
    1. Same format as historical (header row, date, multiple view columns, sum of views)
    2. Simple format: date, views (or total_views)
    3. Simple format: date, views columns without header row
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        return pd.DataFrame()
    
    try:
        # Try reading with header=1 first (historical format)
        try:
            df_raw = pd.read_csv(csv_path, header=1)
            df_raw.columns = df_raw.columns.str.strip()
            has_header = True
        except:
            # Try reading without header offset (simple format)
            df_raw = pd.read_csv(csv_path)
            df_raw.columns = df_raw.columns.str.strip()
            has_header = False
        
        # Find date column
        date_col = None
        for col in df_raw.columns:
            if col.strip().lower() in ["date", "dates"]:
                date_col = col
                break
        
        if date_col is None:
            # Try first column as date
            date_col = df_raw.columns[0]
        
        # Parse dates
        dates = _parse_tiktok_dates(df_raw[date_col])
        
        # Find views column(s)
        sum_of_views_col = None
        for col in df_raw.columns:
            if col.strip().lower() == "sum of views":
                sum_of_views_col = col
                break
        
        if sum_of_views_col:
            # Use "Sum of Views" column if available
            total_views = pd.to_numeric(
                df_raw[sum_of_views_col].astype(str).str.replace(r"[^\d.-]", "", regex=True),
                errors="coerce",
            )
        else:
            # Look for view columns
            view_cols = [col for col in df_raw.columns if "view" in col.lower() and col != date_col]
            
            if view_cols:
                # Sum all view columns
                numeric_cols = df_raw[view_cols].apply(
                    lambda col: pd.to_numeric(
                        col.astype(str).str.replace(r"[^\d.-]", "", regex=True), errors="coerce"
                    )
                )
                total_views = numeric_cols.sum(axis=1, skipna=True)
            else:
                # Try to find a single numeric column (might be total views)
                numeric_cols_found = []
                for col in df_raw.columns:
                    if col != date_col:
                        try:
                            test_vals = pd.to_numeric(
                                df_raw[col].astype(str).str.replace(r"[^\d.-]", "", regex=True),
                                errors="coerce"
                            )
                            if test_vals.notna().any():
                                numeric_cols_found.append(col)
                        except:
                            pass
                
                if numeric_cols_found:
                    # Use the first numeric column found (or sum if multiple)
                    if len(numeric_cols_found) == 1:
                        total_views = pd.to_numeric(
                            df_raw[numeric_cols_found[0]].astype(str).str.replace(r"[^\d.-]", "", regex=True),
                            errors="coerce",
                        )
                    else:
                        # Sum multiple numeric columns
                        numeric_cols = df_raw[numeric_cols_found].apply(
                            lambda col: pd.to_numeric(
                                col.astype(str).str.replace(r"[^\d.-]", "", regex=True), errors="coerce"
                            )
                        )
                        total_views = numeric_cols.sum(axis=1, skipna=True)
                else:
                    total_views = pd.Series([0] * len(df_raw), dtype=float)
        
        df = pd.DataFrame()
        df["date"] = dates
        df["total_views"] = total_views
        
        if "BSR Amazon" not in df.columns:
            df["BSR Amazon"] = pd.NA
        df = df.dropna(subset=["date"])
        df = df[df["total_views"].notna() | df["BSR Amazon"].notna()]
        df = df.sort_values("date")
        return df[["date", "total_views", "BSR Amazon"]]
    except Exception as e:
        # Return empty dataframe on any error
        return pd.DataFrame()


def load_manual_bsr_entries() -> list:
    """Load manual BSR entries from JSON file."""
    if not BSR_MANUAL_ENTRIES_PATH.exists():
        return []
    
    try:
        with open(BSR_MANUAL_ENTRIES_PATH, "r") as f:
            return json.load(f)
    except Exception:
        return []


def save_manual_bsr_entry(date: str, bsr: float) -> bool:
    """Save or update a manual BSR entry."""
    entries = load_manual_bsr_entries()
    
    # Remove existing entry for this date if any
    entries = [e for e in entries if e["date"] != date]
    
    # Add new entry
    entries.append({"date": date, "bsr": float(bsr)})
    
    # Sort by date
    entries.sort(key=lambda x: x["date"])
    
    # Save to file
    try:
        BSR_MANUAL_ENTRIES_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(BSR_MANUAL_ENTRIES_PATH, "w") as f:
            json.dump(entries, f, indent=2)
        return True
    except Exception:
        return False


def delete_manual_bsr_entry(date: str) -> bool:
    """Delete a manual BSR entry."""
    entries = load_manual_bsr_entries()
    entries = [e for e in entries if e["date"] != date]
    
    try:
        with open(BSR_MANUAL_ENTRIES_PATH, "w") as f:
            json.dump(entries, f, indent=2)
        return True
    except Exception:
        return False


def get_file_modification_time(file_path: Path) -> Optional[datetime]:
    """Get file modification time."""
    try:
        return datetime.fromtimestamp(file_path.stat().st_mtime)
    except Exception:
        return None


def parse_google_sheets_url(url: str) -> Optional[dict]:
    """Parse Google Sheets URL to extract sheet ID and GID.
    
    Returns dict with 'sheet_id' and 'gid', or None if URL is invalid.
    """
    # Pattern: https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit?gid={GID}#gid={GID}
    pattern = r'docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)'
    match = re.search(pattern, url)
    
    if not match:
        return None
    
    sheet_id = match.group(1)
    
    # Extract GID from URL (can be in query params or hash)
    gid = None
    if 'gid=' in url:
        gid_match = re.search(r'gid=(\d+)', url)
        if gid_match:
            gid = gid_match.group(1)
    
    return {"sheet_id": sheet_id, "gid": gid}


def get_google_sheets_csv_url(sheet_id: str, gid: Optional[str] = None) -> str:
    """Convert Google Sheets URL to CSV export URL."""
    base_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    if gid:
        base_url += f"&gid={gid}"
    return base_url


@st.cache_data(ttl=60)  # Cache for 1 minute (shorter for live data)
def load_current_views_data_from_google_sheets(url: str) -> pd.DataFrame:
    """Load current mode views data from Google Sheets URL."""
    try:
        # Parse the Google Sheets URL
        parsed = parse_google_sheets_url(url)
        if not parsed:
            return pd.DataFrame()
        # Convert to CSV export URL
        csv_url = get_google_sheets_csv_url(parsed["sheet_id"], parsed.get("gid"))
        
        # Set timeout for network requests
        socket.setdefaulttimeout(10)  # 10 second timeout

        def _has_date_column(frame: pd.DataFrame) -> bool:
            return any(
                isinstance(col, str) and col.strip().lower() in {"date", "dates"}
                for col in frame.columns
            )

        df_raw = pd.DataFrame()
        for header in (4, 0):
            try:
                candidate = pd.read_csv(csv_url, header=header)
                candidate = candidate.dropna(how="all")
                candidate.columns = candidate.columns.map(lambda c: str(c).strip())
                if _has_date_column(candidate):
                    df_raw = candidate
                    break
            except Exception as e:
                logging.warning(f"Error reading CSV with header={header}: {e}")
                continue

        if df_raw.empty:
            try:
                raw = pd.read_csv(csv_url, header=None)
                raw = raw.dropna(how="all")
            except Exception as e:
                logging.warning(f"Error reading CSV without header: {e}")
                return pd.DataFrame()
            header_row_idx = None
            for idx in range(min(15, len(raw))):
                row = raw.iloc[idx].astype(str).str.strip().str.lower()
                if "date" in row.values or "dates" in row.values:
                    header_row_idx = idx
                    break
            if header_row_idx is None:
                return pd.DataFrame()
            header_row = raw.iloc[header_row_idx].astype(str).str.strip()
            df_raw = raw.iloc[header_row_idx + 1 :].reset_index(drop=True)
            df_raw.columns = header_row
            df_raw = df_raw.dropna(how="all")
            df_raw.columns = df_raw.columns.map(lambda c: str(c).strip())
        # Detect date column
        date_col = None
        for col in df_raw.columns:
            if col.strip().lower() in ["date", "dates"]:
                date_col = col
                break
        if date_col is None:
            date_col = df_raw.columns[0]
        # Parse dates and views
        dates = _parse_tiktok_dates(df_raw[date_col])
        views_sum_col = None
        for col in df_raw.columns:
            if col.strip().lower() == "views sum":
                views_sum_col = col
                break
        if views_sum_col:
            total_views = pd.to_numeric(
                df_raw[views_sum_col].astype(str).str.replace(r"[^\d.-]", "", regex=True),
                errors="coerce",
            )
        else:
            view_cols = [col for col in df_raw.columns if "view" in col.lower() and col != date_col]
            if view_cols:
                numeric_cols = df_raw[view_cols].apply(
                    lambda col: pd.to_numeric(
                        col.astype(str).str.replace(r"[^\d.-]", "", regex=True), errors="coerce"
                    )
                )
                total_views = numeric_cols.sum(axis=1, skipna=True)
            else:
                total_views = pd.Series([0] * len(df_raw), dtype=float)
        # Build dataframe and merge manual BSR
        df = pd.DataFrame()
        df["date"] = dates
        df["total_views"] = total_views
        if "BSR Amazon" not in df.columns:
            df["BSR Amazon"] = pd.NA
        df = df.dropna(subset=["date"])
        df = df[df["total_views"].notna() | df["BSR Amazon"].notna()]
        df = df.sort_values("date")
        return df[["date", "total_views", "BSR Amazon"]]
    except Exception as e:
        # Return empty dataframe on any error
        return pd.DataFrame()


def extract_tiktok_urls_from_row(row: pd.Series) -> list[str]:
    """Scan a row for TikTok URLs."""
    urls: list[str] = []
    for value in row:
        if isinstance(value, str):
            matches = re.findall(r"https?://(?:www\\.)?(?:vm\\.)?tiktok\\.com/[^\\s,]+", value)
            urls.extend(matches)
    # Deduplicate while preserving order
    seen = set()
    deduped = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            deduped.append(u)
    return deduped


def parse_tiktok_video_id(url: str) -> Optional[str]:
    """Extract TikTok numeric video id from standard URLs like /video/<id>.
    Returns None for short redirect links where id cannot be parsed.
    """
    try:
        # Normalize
        parsed = urllib.parse.urlparse(url)
        path = parsed.path or ""
        m = re.search(r"/video/(\d+)", path)
        if m:
            return m.group(1)
        return None
    except Exception:
        return None


def parse_tiktok_username(url: str) -> Optional[str]:
    """Extract TikTok username from URL like @username."""
    try:
        parsed = urllib.parse.urlparse(url)
        path = parsed.path or ""
        # Match @username in path like /@username/video/...
        m = re.search(r"/@([^/]+)", path)
        if m:
            return m.group(1)
        return None
    except Exception:
        return None


def get_tiktok_embed_url(video_id: str) -> str:
    """Build TikTok embed URL for a given video id."""
    return f"https://www.tiktok.com/embed/v2/{video_id}"


def get_tiktok_oembed_html(url: str) -> str:
    """Get TikTok embed HTML using the exact structure that works.
    Extracts video ID and username from the URL and builds proper embed.
    """
    logger = logging.getLogger(__name__)
    
    # Extract video ID and username from URL
    video_id = parse_tiktok_video_id(url)
    username = parse_tiktok_username(url)
    
    if not video_id:
        logger.warning(f"Could not extract video ID from {url}, using fallback")
        # Fallback for URLs without video ID
        return f'''<blockquote class="tiktok-embed" cite="{url}" style="max-width: 605px;min-width: 325px;" > <section> <a target="_blank" title="TikTok" href="{url}">View on TikTok</a> </section> </blockquote> <script async src="https://www.tiktok.com/embed.js"></script>'''
    
    # Build username link - use extracted username or fallback
    if username:
        username_link = f'<a target="_blank" title="@{username}" href="https://www.tiktok.com/@{username}?refer=embed">@{username}</a>'
    else:
        # Try to extract from URL path as fallback
        try:
            parsed = urllib.parse.urlparse(url)
            path_parts = [p for p in parsed.path.split('/') if p]
            for i, part in enumerate(path_parts):
                if part.startswith('@'):
                    username = part[1:]  # Remove @
                    username_link = f'<a target="_blank" title="@{username}" href="https://www.tiktok.com/@{username}?refer=embed">@{username}</a>'
                    break
            else:
                username_link = f'<a target="_blank" title="TikTok" href="{url}">TikTok</a>'
        except:
            username_link = f'<a target="_blank" title="TikTok" href="{url}">TikTok</a>'
    
    # Build embed HTML using the exact structure provided
    embed_html = f'''<blockquote class="tiktok-embed" cite="{url}" data-video-id="{video_id}" style="max-width: 605px;min-width: 325px;" > <section> {username_link} <p></p> </section> </blockquote> <script async src="https://www.tiktok.com/embed.js"></script>'''
    
    return embed_html


@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_video_details_long(csv_path: Path = VIDEO_DETAILS_DATA_PATH) -> pd.DataFrame:
    """Load the video-details export and reshape it into a long dataframe."""
    csv_path = Path(csv_path)
    if not csv_path.exists():
        return pd.DataFrame()

    try:
        df_raw = pd.read_csv(csv_path)
    except Exception:
        return pd.DataFrame()

    if df_raw.empty:
        return pd.DataFrame()

    meta_cols = ["Video ID", "Type", "Video URL", "Created Date"]
    if not set(meta_cols).issubset(df_raw.columns):
        return pd.DataFrame()

    df_meta = (
        df_raw[meta_cols]
        .rename(
            columns={
                "Video ID": "video_id",
                "Type": "video_type",
                "Video URL": "video_url",
                "Created Date": "created_date",
            }
        )
        .reset_index(drop=True)
    )
    df_meta["created_date"] = _parse_tiktok_dates(df_meta["created_date"])

    metric_cols = [col for col in df_raw.columns if col not in meta_cols]
    column_groups: dict[tuple[str, str], list[str]] = {}
    for col in metric_cols:
        if not isinstance(col, str) or " " not in col:
            continue
        try:
            date_part, metric = col.rsplit(" ", 1)
        except ValueError:
            continue
        metric = metric.strip().lower()
        if metric not in {"views", "likes", "comments", "shares"}:
            continue
        date_part = date_part.strip()
        column_groups.setdefault((date_part, metric), []).append(col)

    if not column_groups:
        return pd.DataFrame()

    def _date_sort_key(label: str):
        parsed = pd.to_datetime(label, errors="coerce")
        return parsed if pd.notna(parsed) else pd.Timestamp.max

    ordered_dates = sorted({date for date, _ in column_groups.keys()}, key=_date_sort_key)
    ordered_data: dict[tuple[str, str], pd.Series] = {}
    for date_label in ordered_dates:
        for metric in ("views", "likes", "comments", "shares"):
            cols = column_groups.get((date_label, metric))
            if not cols:
                continue
            numeric_values = df_raw[cols].apply(pd.to_numeric, errors="coerce")
            aggregated = numeric_values.max(axis=1, skipna=True)
            ordered_data[(date_label, metric)] = aggregated

    if not ordered_data:
        return pd.DataFrame()

    wide_df = pd.DataFrame(ordered_data)
    wide_df.columns = pd.MultiIndex.from_tuples(wide_df.columns, names=["date_label", "metric"])
    wide_df["row_index"] = wide_df.index
    stacked = wide_df.set_index("row_index").stack(level="date_label").reset_index()
    stacked = stacked.rename(columns={"date_label": "date"})
    metric_columns = [col for col in stacked.columns if col not in {"row_index", "date"}]
    stacked = stacked.rename(columns={col: col.lower() for col in metric_columns})

    meta_aligned = df_meta.loc[stacked["row_index"]].reset_index(drop=True)
    long_df = pd.concat([meta_aligned, stacked.drop(columns="row_index").reset_index(drop=True)], axis=1)
    long_df["date"] = _parse_tiktok_dates(long_df["date"])
    for col in ["views", "likes", "comments", "shares"]:
        if col in long_df.columns:
            long_df[col] = pd.to_numeric(long_df[col], errors="coerce")
    long_df = long_df.dropna(subset=["date"])
    return long_df


def summarize_video_details(details_df: pd.DataFrame) -> pd.DataFrame:
    """Summarize per-video performance over the observed window."""
    if details_df.empty:
        return pd.DataFrame()

    sorted_df = details_df.sort_values(["video_id", "date"])
    metrics = sorted_df.groupby(
        ["video_id", "video_type", "video_url", "created_date"],
        as_index=False,
    ).agg(
        first_date=("date", "first"),
        last_date=("date", "last"),
        first_views=("views", "first"),
        latest_views=("views", "last"),
        latest_likes=("likes", "last"),
        latest_comments=("comments", "last"),
        latest_shares=("shares", "last"),
    )

    metrics["views_delta"] = metrics["latest_views"] - metrics["first_views"]
    metrics["views_growth_pct"] = (
        metrics["views_delta"] / metrics["first_views"].replace(0, pd.NA)
    ) * 100
    metrics["observation_days"] = (metrics["last_date"] - metrics["first_date"]).dt.days + 1
    metrics["days_since_created"] = (metrics["last_date"] - metrics["created_date"]).dt.days
    
    # Calculate average daily increment: total change divided by number of intervals
    # If we observe 7 days, the change happens over 6 intervals (day 1->2, 2->3, ..., 6->7)
    # So we divide by (observation_days - 1) to get the average daily increment
    intervals = (metrics["observation_days"] - 1).replace(0, pd.NA)
    metrics["avg_daily_views"] = metrics["views_delta"] / intervals
    metrics["avg_daily_views"] = metrics["avg_daily_views"].fillna(0)

    return metrics


@st.cache_data(ttl=300)  # Cache for 5 minutes
def get_date_to_tiktok_urls_from_google_sheets(url: str, sheet_name: Optional[str] = None) -> dict:
    """Return mapping of date (pd.Timestamp normalized to day) -> list of TikTok URLs found on that row.
    IMPORTANT: We use the XLSX export so we can read real cell hyperlinks (CSV loses links).
    """
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    mapping: dict = {}
    try:
        parsed = parse_google_sheets_url(url)
        if not parsed:
            logger.warning("Failed to parse Google Sheets URL")
            return mapping
        logger.info(f"Parsed sheet_id: {parsed['sheet_id']}, gid: {parsed.get('gid')}")
        
        # Download XLSX with timeout
        xlsx_url = f"https://docs.google.com/spreadsheets/d/{parsed['sheet_id']}/export?format=xlsx"
        logger.info(f"Downloading XLSX from: {xlsx_url}")
        socket.setdefaulttimeout(15)  # 15 second timeout for XLSX download
        try:
            with urllib.request.urlopen(xlsx_url, timeout=15) as resp:
                content = resp.read()
            logger.info(f"Downloaded {len(content)} bytes")
        except Exception as e:
            logger.error(f"Error downloading XLSX: {e}")
            return mapping
        finally:
            socket.setdefaulttimeout(None)
        
        # IMPORTANT: Don't use data_only=True - it strips hyperlinks!
        # Try with keep_links=True if available (openpyxl 3.1+)
        try:
            wb = load_workbook(BytesIO(content), data_only=False, keep_links=True)
        except TypeError:
            # Older openpyxl versions don't have keep_links
            wb = load_workbook(BytesIO(content), data_only=False)
        logger.info(f"Workbook loaded. Sheet names: {wb.sheetnames}")

        # If a specific sheet is requested, try it first
        candidate_sheets = []
        if sheet_name:
            logger.info(f"Looking for sheet: {sheet_name}")
            if sheet_name in wb.sheetnames:
                candidate_sheets.append(wb[sheet_name])
                logger.info(f"Found requested sheet: {sheet_name}")
            else:
                logger.warning(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        # Add remaining sheets
        for ws in wb.worksheets:
            if ws not in candidate_sheets:
                candidate_sheets.append(ws)
        logger.info(f"Processing {len(candidate_sheets)} sheet(s)")

        # Heuristics: pick the first sheet (or specified one) that contains dates
        for ws in candidate_sheets:
            logger.info(f"Processing sheet: {ws.title}, max_row={ws.max_row}, max_column={ws.max_column}")
            date_col_idx = None
            header_row_idx = None
            
            # First, try to find explicit "Date" header
            for r in range(1, min(ws.max_row, 20) + 1):
                row_vals = [str(ws.cell(r, c).value).strip().lower() if ws.cell(r, c).value is not None else "" for c in range(1, min(ws.max_column, 60) + 1)]
                if any(val == "date" or val == "dates" for val in row_vals):
                    header_row_idx = r
                    for c, val in enumerate(row_vals, start=1):
                        if val in ("date", "dates"):
                            date_col_idx = c
                            break
                    break
            
            # If no explicit Date header, assume dates are in column A (first column)
            if header_row_idx is None or date_col_idx is None:
                logger.info(f"No explicit 'Date' header found, assuming dates in column A")
                date_col_idx = 1  # Column A
                # Try to find header row by looking for common headers like "Views", "Likes", etc.
                for r in range(1, min(ws.max_row, 10) + 1):
                    row_vals = [str(ws.cell(r, c).value).strip().lower() if ws.cell(r, c).value is not None else "" for c in range(1, min(ws.max_column, 60) + 1)]
                    if any(val in ("views", "likes", "comments", "shares") for val in row_vals):
                        header_row_idx = r
                        break
                # If still no header row found, assume data starts at row 2
                if header_row_idx is None:
                    header_row_idx = 1
                    logger.info(f"No header row found, assuming data starts at row {header_row_idx + 1}")
            
            logger.info(f"Using Date column: {date_col_idx}, Header row: {header_row_idx}")

            # Iterate rows after header, collect date + any TikTok hyperlinks in the row
            rows_processed = 0
            hyperlinks_found = 0
            total_cells_checked = 0
            
            # First, let's check what hyperlinks exist in the sheet
            logger.info(f"Checking for hyperlinks in sheet {ws.title}...")
            all_hyperlinks = []
            all_formulas = []
            for r in range(1, min(ws.max_row + 1, 50)):
                for c in range(1, min(ws.max_column + 1, 30)):
                    try:
                        cell = ws.cell(r, c)
                        # Check hyperlink attribute
                        hl = getattr(cell, "hyperlink", None)
                        if hl:
                            target = getattr(hl, "target", None)
                            if target:
                                all_hyperlinks.append((r, c, target))
                                logger.info(f"Found hyperlink at Row {r}, Col {chr(64+c)} ({c}): {target}")
                        # Check formula
                        if hasattr(cell, 'formula') and cell.formula:
                            formula = str(cell.formula)
                            if "HYPERLINK" in formula.upper():
                                all_formulas.append((r, c, formula))
                                logger.info(f"Found HYPERLINK formula at Row {r}, Col {chr(64+c)} ({c}): {formula}")
                    except Exception as e:
                        logger.debug(f"Error checking cell {r},{c}: {e}")
                        pass
            
            logger.info(f"Total hyperlinks found in first 50 rows: {len(all_hyperlinks)}")
            logger.info(f"Total HYPERLINK formulas found in first 50 rows: {len(all_formulas)}")
            
            # NEW APPROACH: Each Views column represents ONE video
            # Map: column_index -> TikTok URL (extracted from that column)
            column_to_url: dict[int, str] = {}
            
            # First, identify which columns are "Views" columns
            views_columns = []
            for c in range(1, ws.max_column + 1):
                try:
                    # Check header row for "Views" label
                    header_cell = ws.cell(header_row_idx, c)
                    header_val = str(header_cell.value).strip().lower() if header_cell.value else ""
                    if "view" in header_val and "sum" not in header_val:
                        views_columns.append(c)
                        logger.info(f"Found Views column at column {chr(64+c)} ({c})")
                except:
                    pass
            
            # For each Views column, extract the TikTok URL from the first data row
            for col_idx in views_columns:
                url_found = None
                # Check first few data rows for hyperlink in this column
                for r in range(header_row_idx + 1, min(header_row_idx + 10, ws.max_row + 1)):
                    try:
                        cell = ws.cell(r, col_idx)
                        
                        # Check hyperlink attribute
                        hl = getattr(cell, "hyperlink", None)
                        if hl:
                            target = getattr(hl, "target", None)
                            if target and "tiktok.com" in target.lower():
                                url_found = target
                                logger.info(f"Column {chr(64+col_idx)} ({col_idx}): Found TikTok URL in row {r}: {url_found}")
                                break
                        
                        # Check formula
                        if hasattr(cell, 'formula') and cell.formula:
                            formula = str(cell.formula)
                            if "HYPERLINK" in formula.upper():
                                match = re.search(r'HYPERLINK\(["\']([^"\']+)["\']', formula, re.IGNORECASE)
                                if match:
                                    url = match.group(1)
                                    if "tiktok.com" in url.lower():
                                        url_found = url
                                        logger.info(f"Column {chr(64+col_idx)} ({col_idx}): Found TikTok URL via formula in row {r}: {url_found}")
                                        break
                        
                        # Check cell value
                        cell_val = cell.value
                        if cell_val and isinstance(cell_val, str) and "tiktok.com" in cell_val.lower():
                            url_found = cell_val
                            logger.info(f"Column {chr(64+col_idx)} ({col_idx}): Found TikTok URL in cell value row {r}: {url_found}")
                            break
                            
                    except Exception as e:
                        logger.debug(f"Error checking column {col_idx}, row {r}: {e}")
                        continue
                
                if url_found:
                    column_to_url[col_idx] = url_found
                    logger.info(f"Mapped column {chr(64+col_idx)} ({col_idx}) -> {url_found}")
            
            logger.info(f"Found {len(column_to_url)} video columns with TikTok URLs")
            
            # Now iterate through data rows and map dates to videos based on which columns have data
            for r in range(header_row_idx + 1, ws.max_row + 1):
                date_cell = ws.cell(r, date_col_idx)
                date_val = date_cell.value
                if date_val is None or str(date_val).strip() == "":
                    continue
                
                # Parse the date
                try:
                    if isinstance(date_val, (pd.Timestamp, datetime, date)):
                        day = pd.to_datetime(date_val).normalize()
                    else:
                        parsed_date = _parse_tiktok_dates(pd.Series([date_val])).iloc[0]
                        if pd.isna(parsed_date):
                            continue
                        day = pd.to_datetime(parsed_date).normalize()
                    rows_processed += 1
                except Exception as e:
                    continue

                # For this date, find which Views columns have non-zero values
                urls_for_date = []
                for col_idx, url in column_to_url.items():
                    try:
                        cell = ws.cell(r, col_idx)
                        cell_val = cell.value
                        # Check if this cell has a value (views > 0 or has hyperlink)
                        if cell_val is not None:
                            # Try to parse as number
                            try:
                                views = float(cell_val)
                                if views > 0:
                                    urls_for_date.append(url)
                            except:
                                # Not a number, but cell has content - include it
                                if str(cell_val).strip():
                                    urls_for_date.append(url)
                    except:
                        pass

                if urls_for_date:
                    existing = mapping.get(day, [])
                    for u in urls_for_date:
                        if u not in existing:
                            existing.append(u)
                    mapping[day] = existing
                    logger.debug(f"Date {day}: Added {len(urls_for_date)} TikTok URL(s) from {len(column_to_url)} video columns")
            
            logger.info(f"Sheet {ws.title}: Processed {rows_processed} rows, checked {total_cells_checked} cells, found {hyperlinks_found} TikTok hyperlinks")

            # If we populated mapping from this sheet, we can stop
            if mapping:
                logger.info(f"Found {len(mapping)} dates with TikTok links in sheet {ws.title}")
                break

        logger.info(f"Final mapping: {len(mapping)} dates with TikTok URLs")
        return mapping
    except Exception as e:
        logger.error(f"Error extracting TikTok URLs: {e}", exc_info=True)
        return mapping


# ---------- Manual TikTok links storage ----------
def load_manual_tiktok_links() -> dict:
    """Load manual TikTok links mapping: { 'YYYY-MM-DD': [url, ...] }"""
    try:
        if not MANUAL_TIKTOK_LINKS_PATH.exists():
            return {}
        with open(MANUAL_TIKTOK_LINKS_PATH, "r") as f:
            data = json.load(f)
            if isinstance(data, dict):
                return data
            return {}
    except Exception:
        return {}


def save_manual_tiktok_links(date_str: str, urls: list[str]) -> bool:
    """Save or merge manual TikTok links for a specific date (YYYY-MM-DD)."""
    try:
        data = load_manual_tiktok_links()
        existing = data.get(date_str, [])
        # Merge unique while preserving order
        seen = set(existing)
        for u in urls:
            if u and u not in seen:
                existing.append(u)
                seen.add(u)
        data[date_str] = existing
        MANUAL_TIKTOK_LINKS_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(MANUAL_TIKTOK_LINKS_PATH, "w") as f:
            json.dump(data, f, indent=2)
        return True
    except Exception:
        return False


def merge_with_manual_tiktok_links(date_to_urls: dict) -> dict:
    """Merge Google Sheets URL mapping with manually-added links."""
    merged = {**date_to_urls}
    manual = load_manual_tiktok_links()
    for date_str, urls in manual.items():
        try:
            day = pd.to_datetime(date_str).normalize()
        except Exception:
            continue
        current = merged.get(day, [])
        for u in urls:
            if u not in current:
                current.append(u)
        merged[day] = current
    return merged


# ---------- Durable BSR persistence via SQLite ----------
def _init_db():
    """Initialize SQLite database and migrate JSON BSR entries if present."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(DB_PATH) as conn:
            # Check if table exists and what columns it has
            cursor = conn.execute("PRAGMA table_info(manual_bsr)")
            columns = [row[1] for row in cursor.fetchall()]
            
            if not columns:
                # Table doesn't exist, create with brand support
                conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS manual_bsr (
                        date TEXT,
                        brand TEXT DEFAULT 'Trueseamoss',
                        bsr REAL,
                        PRIMARY KEY (date, brand)
                    )
                    """
                )
            elif "brand" not in columns:
                # Table exists but doesn't have brand column - migrate
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS manual_bsr_new (
                        date TEXT,
                        brand TEXT DEFAULT 'Trueseamoss',
                        bsr REAL,
                        PRIMARY KEY (date, brand)
                    )
                """)
                conn.execute("""
                    INSERT INTO manual_bsr_new (date, brand, bsr)
                    SELECT date, 'Trueseamoss', bsr FROM manual_bsr
                """)
                conn.execute("DROP TABLE manual_bsr")
                conn.execute("ALTER TABLE manual_bsr_new RENAME TO manual_bsr")
                conn.commit()
            else:
                # Table exists with brand column, just ensure it exists
                conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS manual_bsr (
                        date TEXT,
                        brand TEXT DEFAULT 'Trueseamoss',
                        bsr REAL,
                        PRIMARY KEY (date, brand)
                    )
                    """
                )
            
            # Always ensure initial BSR data exists with correct values (use INSERT OR REPLACE)
            initial_bsr_data = [
                # Trueseamoss BSR data
                ("2025-11-30", "Trueseamoss", 86.0),
                ("2025-12-01", "Trueseamoss", 84.0),
                ("2025-12-02", "Trueseamoss", 90.0),
                ("2025-12-03", "Trueseamoss", 92.0),
                # HerbalVineyard BSR data
                ("2025-12-01", "HerbalVineyard", 22201.0),
                ("2025-12-02", "HerbalVineyard", 22506.0),
                ("2025-12-03", "HerbalVineyard", 19050.0),
            ]
            for date, brand, bsr in initial_bsr_data:
                # Use INSERT OR REPLACE to ensure correct values (overwrites if exists)
                conn.execute(
                    "INSERT OR REPLACE INTO manual_bsr(date, brand, bsr) VALUES(?, ?, ?)",
                    (date, brand, bsr)
                )
            conn.commit()
            
            # Migrate from JSON once if table was empty and JSON exists (legacy support)
            cur = conn.execute("SELECT COUNT(*) FROM manual_bsr")
            count = cur.fetchone()[0] or 0
            if count == len(initial_bsr_data) and BSR_MANUAL_ENTRIES_PATH.exists():
                try:
                    with open(BSR_MANUAL_ENTRIES_PATH, "r") as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            for entry in data:
                                d = entry.get("date")
                                b = entry.get("bsr")
                                if d is not None and b is not None:
                                    conn.execute(
                                        "INSERT OR REPLACE INTO manual_bsr(date, brand, bsr) VALUES(?, ?, ?)",
                                        (str(d), "Trueseamoss", float(b)),
                                    )
                            conn.commit()
                except Exception:
                    pass
    except Exception:
        pass


@st.cache_data(ttl=10)  # Cache for 10 seconds (shorter to pick up new BSR entries faster)
def load_manual_bsr_entries(brand: str = "Trueseamoss") -> list:
    """Load manual BSR entries from SQLite for a specific brand (migrates JSON if needed)."""
    _init_db()
    try:
        with sqlite3.connect(DB_PATH) as conn:
            rows = conn.execute(
                "SELECT date, bsr FROM manual_bsr WHERE brand = ? ORDER BY date",
                (brand,)
            ).fetchall()
            return [{"date": r[0], "bsr": float(r[1]) if r[1] is not None else None} for r in rows]
    except Exception:
        # Fallback to JSON (legacy) - only for Trueseamoss
        if brand != "Trueseamoss" or not BSR_MANUAL_ENTRIES_PATH.exists():
            return []
        try:
            with open(BSR_MANUAL_ENTRIES_PATH, "r") as f:
                return json.load(f)
        except Exception:
            return []


def save_manual_bsr_entry(date: str, bsr: float, brand: str = "Trueseamoss") -> bool:
    """Upsert a manual BSR entry into SQLite for a specific brand."""
    _init_db()
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "INSERT OR REPLACE INTO manual_bsr(date, brand, bsr) VALUES(?, ?, ?)",
                (str(date), brand, float(bsr)),
            )
            conn.commit()
        # Clear cache to reflect new data
        load_manual_bsr_entries.clear()
        return True
    except Exception:
        return False


def delete_manual_bsr_entry(date: str, brand: str = "Trueseamoss") -> bool:
    """Delete a manual BSR entry from SQLite for a specific brand."""
    _init_db()
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("DELETE FROM manual_bsr WHERE date = ? AND brand = ?", (str(date), brand))
            conn.commit()
        # Clear cache to reflect deleted data
        load_manual_bsr_entries.clear()
        return True
    except Exception:
        return False